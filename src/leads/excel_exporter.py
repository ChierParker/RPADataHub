"""Excel export helpers for scraper results."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import LOG_DIR, OUTPUT_COLUMNS
from lead_processing import sanitize_sheet_name
from logger_config import TraceLogger


_logger = TraceLogger("LeadScraper.Excel", str(LOG_DIR))


def save_target_sheet(sheet_name: str, leads: list[dict], output_path: Path) -> None:
    """Write one target's leads into a dedicated Excel sheet."""
    sheet_name = sanitize_sheet_name(sheet_name)
    if not leads:
        return

    df = pd.DataFrame(leads)
    for column in OUTPUT_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df = df[OUTPUT_COLUMNS]
    df[OUTPUT_COLUMNS[0]] = range(1, len(df) + 1)

    if output_path.exists():
        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    format_excel_sheet(output_path, sheet_name)
    _logger.info(f"Sheet [{sheet_name}] wrote {len(df)} rows -> {output_path.name}")


def format_excel_sheet(filepath: Path, sheet_name: str) -> None:
    """Apply basic readability formatting to an exported sheet."""
    try:
        workbook = load_workbook(filepath)
        if sheet_name not in workbook.sheetnames:
            return

        worksheet = workbook[sheet_name]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
            column_letter = get_column_letter(column_index)
            max_width = 10
            for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, min(len(str(cell.value)), 50))
            worksheet.column_dimensions[column_letter].width = max_width + 4

        worksheet.freeze_panes = "A2"
        workbook.save(filepath)
    except Exception as exc:
        _logger.warning(f"Failed to format Excel sheet: {exc}")
